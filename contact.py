from openpyxl import *
from openpyxl.utils import get_column_letter


class Modify:
    def __init__(self, filename, title=''):
        self.n_rows = None
        self.wb = None
        self.sheet = None
        self.person_data = None

        self.labels = ['NO.', 'LAST NAME', 'FIRST NAME', 'CONTACT NUMBER', 'EMAIL', 'HOME ADDRESS']
        self.filename = filename
        self.title = title

    def create(self, data):
        self.wb = Workbook()
        self.sheet = self.wb.active
        if self.title != '':
            self.sheet.title = self.title
        self.sheet.append(self.labels)
        for index, dict in enumerate(data):
            self.sheet.append(list(dict.values()))
        self.wb.save(self.filename)

    def view(self):
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb.active
        n_rows = self.sheet.max_row
        data = []
        for row in range(2, n_rows + 1):
            person = []
            for col in range(1, 7):
                char = get_column_letter(col)
                value = self.sheet[char + str(row)].value
                person.append(value)
            d = dict(zip(self.labels, person))
            data.append(d)
        return data

    def search(self, key):
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb.active
        n_rows = self.sheet.max_row
        data = []
        for row in range(2, n_rows + 1):
            person = []
            for col in range(1, 7):
                char = get_column_letter(col)
                value = self.sheet[char + str(row)].value
                person.append(value)
            if set(key).issubset(set(person)):
                data.append(person)
        return data

    def add(self, data):
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb.active
        for index, dict in enumerate(data):
            self.sheet.append(list(dict.values()))
        self.wb.save(self.filename)

    def update(self, num, label, new_data):
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb.active
        n_rows = self.sheet.max_row
        label_col = get_column_letter(self.labels.index(label)+1)
        for row in range(2, n_rows + 1):
            value = int(self.sheet['A' + str(row)].value)
            if value == num:
                self.sheet[label_col + str(row)].value = new_data
                break
        self.wb.save(self.filename)

    def delete_all(self):
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb.active
        n_rows = self.sheet.max_row
        self.sheet.delete_rows(2, n_rows)
        self.wb.save(self.filename)

    def delete(self, num):
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb.active
        n_rows = self.sheet.max_row
        for row in range(2, n_rows + 1):
            value = int(self.sheet['A' + str(row)].value)
            if value == num:
                self.sheet.delete_rows(row)
                break
        self.wb.save(self.filename)


    def get_npersons(self):
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb.active
        return self.sheet.max_row - 1
